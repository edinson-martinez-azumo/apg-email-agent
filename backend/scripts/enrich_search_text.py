"""
Enrich search_text for products with poor/empty descriptions using Claude few-shot.
Groups products by type, uses well-enriched same-type products as few-shot examples.
Writes back to products_unified.xlsx and re-embeds enriched SKUs (local DB only).

Usage (from backend/):
    uv run python scripts/enrich_search_text.py            # dry run, print stats only
    uv run python scripts/enrich_search_text.py --write    # update xlsx only
    uv run python scripts/enrich_search_text.py --write --embed  # update xlsx + re-embed
    uv run python scripts/enrich_search_text.py --write --embed --limit=50  # first N SKUs
"""
import asyncio
import json
import os
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import anthropic
import cohere
import httpx
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import NullPool

from app.core.config import settings
from app.db.models.product_embedding_v2 import ProductEmbeddingV2
from app.services.product_service import _load_products

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'products_unified.xlsx')
SHORT_THRESHOLD = 150   # chars — below this = needs enrichment
GOOD_THRESHOLD = 300    # chars — above this = usable as few-shot example
BATCH_SIZE = 10         # SKUs per Claude call
EMBED_MODEL = 'embed-english-v3.0'
EMBED_BATCH = 96
EMBED_SLEEP = 8.0

ENRICH_PROMPT = """\
You are enriching product catalog entries for APG (A Packaging Group), a cosmetic packaging supplier.

Generate a rich search_text for each product listed below. The format must be:
{{title}} | {{product_type}} | {{materials}} | {{shopify_tags}} | {{capacity_or_output}}

Shopify tags must include relevant items from these categories (comma-separated):
- Application: Beauty, Beauty Products, Body Wash Products, Hair Products, Skincare Products, Face Cream, Facial Care, Essential Oils, Perfume, Nail Products, Cleaning Products, etc.
- Product type tag matching the product (e.g. Airless Pump Bottle, Lotion Pump, Fine Mist Sprayer, Jar, etc.)
- Material tags: Material_PP, Material_PET, Material_ABS, Material_Glass, Material_Aluminum, Material_PCR, etc.
- Capacity tags if relevant: Capacity_30ml, Capacity_50ml, Capacity_100ml, etc.
- Feature tags: PCR, Recyclable, Airless, Refillable, Travel Size, etc. (only if applicable)

--- EXAMPLES FOR TYPE: {product_type} ---
{examples}
--- END EXAMPLES ---

Now generate search_text for each product below. Return ONLY a JSON array:
[{{"sku": "...", "search_text": "..."}}, ...]

Products to enrich:
{products}
"""


def _clean(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s.lower() in ('nan', 'none', 'false', '') else s


def _st_len(row) -> int:
    return len(_clean(row.get('search_text')))


def _build_examples(df_good: pd.DataFrame, product_type: str, n: int = 3) -> str:
    """Return n few-shot example strings from well-enriched SKUs of the same type."""
    same = df_good[df_good['type'].fillna('') == product_type]
    if len(same) < n:
        # Fallback: any well-enriched SKU
        same = df_good
    sample = same.sample(min(n, len(same)), random_state=42)
    lines = []
    for _, r in sample.iterrows():
        lines.append(
            f"SKU: {_clean(r['sku'])}\n"
            f"Title: {_clean(r['title'])}\n"
            f"Type: {_clean(r['type'])}\n"
            f"Materials: {_clean(r['materials'])}\n"
            f"Capacities: {_clean(r['capacities'])}\n"
            f"search_text: {_clean(r['search_text'])}"
        )
    return '\n\n'.join(lines)


def _build_product_lines(batch: list[dict]) -> str:
    lines = []
    for i, r in enumerate(batch, 1):
        lines.append(
            f"{i}. SKU: {_clean(r['sku'])} | "
            f"Title: {_clean(r['title'])} | "
            f"Type: {_clean(r['type'])} | "
            f"Materials: {_clean(r['materials'])} | "
            f"Capacities: {_clean(r['capacities'])} | "
            f"MOQ: {_clean(r['moq'])}"
        )
    return '\n'.join(lines)


async def _enrich_batch(
    client: anthropic.AsyncAnthropic,
    batch: list[dict],
    examples_str: str,
    product_type: str,
) -> dict[str, str]:
    """Call Claude Haiku for a batch of products. Returns {sku: search_text}."""
    prompt = ENRICH_PROMPT.format(
        product_type=product_type or 'Cosmetic Packaging',
        examples=examples_str,
        products=_build_product_lines(batch),
    )
    for attempt in range(4):
        try:
            resp = await client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=2048,
                temperature=0,
                messages=[{'role': 'user', 'content': prompt}],
            )
            raw = resp.content[0].text.strip()
            # Strip markdown fences
            if raw.startswith('```'):
                raw = raw.split('```', 2)[1]
                if raw.startswith('json'):
                    raw = raw[4:]
            results = json.loads(raw.strip())
            return {item['sku']: item['search_text'] for item in results if 'sku' in item and 'search_text' in item}
        except Exception as e:
            if attempt == 3:
                print(f'\n  [enrich_batch] failed after 4 attempts: {e}')
                return {}
            await asyncio.sleep(5 * (2 ** attempt))
    return {}


async def _reembed_skus(skus: list[str], df: pd.DataFrame) -> None:
    """Re-embed a specific set of SKUs using their updated search_text from df."""
    db_url = settings.database_url
    engine = create_async_engine(db_url, poolclass=NullPool)
    async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    cohere_key = settings.cohere_api_key
    client = cohere.AsyncClientV2(api_key=cohere_key.strip())

    records = df[df['sku'].isin(skus)].to_dict('records')
    print(f'\nRe-embedding {len(records)} enriched SKUs...')

    for i in range(0, len(records), EMBED_BATCH):
        batch = records[i:i + EMBED_BATCH]
        texts = [_clean(r.get('search_text')) or _clean(r.get('title')) or _clean(r.get('sku')) for r in batch]

        for attempt in range(5):
            try:
                resp = await client.embed(
                    texts=texts,
                    model=EMBED_MODEL,
                    input_type='search_document',
                    embedding_types=['float'],
                )
                break
            except (httpx.ReadTimeout, cohere.errors.too_many_requests_error.TooManyRequestsError) as e:
                wait = 15 * (2 ** attempt)
                print(f'\n  [{type(e).__name__}] retry {attempt+1}/5 in {wait}s...')
                await asyncio.sleep(wait)
        else:
            print(f'\nEmbed batch {i // EMBED_BATCH + 1} failed. Skipping.')
            continue

        async with async_session() as session:
            for rec, embedding in zip(batch, resp.embeddings.float_):
                search_text = _clean(rec.get('search_text')) or _clean(rec.get('title'))
                stmt = pg_insert(ProductEmbeddingV2).values(
                    sku=_clean(rec['sku']),
                    title=_clean(rec.get('title')) or None,
                    search_text=search_text or None,
                    embedding=embedding,
                    type=_clean(rec.get('type')) or None,
                    materials=_clean(rec.get('materials')) or None,
                    moq=_clean(rec.get('moq')) or None,
                    capacities=_clean(rec.get('capacities')) or None,
                    price_base=_clean(rec.get('price_base')) or None,
                    price_10k=_clean(rec.get('price_10k')) or None,
                    price_25k=_clean(rec.get('price_25k')) or None,
                    price_50k=_clean(rec.get('price_50k')) or None,
                    price_100k=_clean(rec.get('price_100k')) or None,
                    in_stock=bool(rec.get('in_stock', False)),
                    image_url=_clean(rec.get('image_url')) or None,
                    dimensions=_clean(rec.get('dimensions')) or None,
                    search_vector=search_text or None,
                ).on_conflict_do_update(
                    index_elements=['sku'],
                    set_=dict(
                        title=_clean(rec.get('title')) or None,
                        search_text=search_text or None,
                        embedding=embedding,
                        type=_clean(rec.get('type')) or None,
                        materials=_clean(rec.get('materials')) or None,
                        moq=_clean(rec.get('moq')) or None,
                        capacities=_clean(rec.get('capacities')) or None,
                        price_base=_clean(rec.get('price_base')) or None,
                        price_10k=_clean(rec.get('price_10k')) or None,
                        price_25k=_clean(rec.get('price_25k')) or None,
                        price_50k=_clean(rec.get('price_50k')) or None,
                        price_100k=_clean(rec.get('price_100k')) or None,
                        in_stock=bool(rec.get('in_stock', False)),
                        image_url=_clean(rec.get('image_url')) or None,
                        dimensions=_clean(rec.get('dimensions')) or None,
                        search_vector=search_text or None,
                    ),
                )
                await session.execute(stmt)
            await session.commit()

        done = min(i + EMBED_BATCH, len(records))
        print(f'  embedded {done}/{len(records)}', end='\r')

        if i + EMBED_BATCH < len(records):
            await asyncio.sleep(EMBED_SLEEP)

    print(f'\nEmbed complete: {len(records)} SKUs updated.')
    await engine.dispose()


def _write_xlsx(df: pd.DataFrame, enriched_skus: list[str]) -> None:
    """Write updated search_text back to xlsx using openpyxl to preserve formatting."""
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    # Find column indices
    header = {cell.value: cell.column for cell in ws[1]}
    sku_col = header.get('sku')
    st_col = header.get('search_text')

    if not sku_col or not st_col:
        raise ValueError(f'Missing sku or search_text column in xlsx. Headers: {list(header.keys())}')

    sku_to_st = {str(r['sku']): str(r['search_text']) for _, r in df[df['sku'].isin(enriched_skus)].iterrows()}

    updated = 0
    for row in ws.iter_rows(min_row=2):
        sku_cell = row[sku_col - 1]
        st_cell = row[st_col - 1]
        sku_val = str(sku_cell.value or '').strip()
        if sku_val in sku_to_st:
            st_cell.value = sku_to_st[sku_val]
            updated += 1

    wb.save(XLSX_PATH)
    print(f'Wrote {updated} updated search_text values to {os.path.basename(XLSX_PATH)}')


async def run(write: bool, embed: bool, limit: int | None) -> None:
    df = _load_products()
    df = df.drop_duplicates('sku')

    # Identify SKUs needing enrichment
    needs = df[df.apply(_st_len, axis=1) < SHORT_THRESHOLD].copy()
    if limit:
        needs = needs.head(limit)

    print(f'SKUs needing enrichment (search_text < {SHORT_THRESHOLD} chars): {len(needs)}')

    # Well-enriched pool for few-shot examples
    good = df[df.apply(_st_len, axis=1) >= GOOD_THRESHOLD].copy()
    print(f'Available few-shot examples (search_text >= {GOOD_THRESHOLD} chars): {len(good)}')

    if not write:
        print('\nDry run — pass --write to update xlsx, --write --embed to also re-embed.')
        # Show type breakdown
        print('\nBreakdown by type:')
        print(needs['type'].fillna('(no type)').value_counts().head(15).to_string())
        return

    # Group by type and enrich
    anthropic_client = anthropic.AsyncAnthropic(api_key=settings.anthropic_api_key)
    enriched: dict[str, str] = {}  # sku -> new search_text

    type_groups = needs.groupby(needs['type'].fillna('(no type)'))
    n_types = len(type_groups)
    t_start = time.time()

    for t_idx, (product_type, group) in enumerate(type_groups, 1):
        rows = group.to_dict('records')
        examples_str = _build_examples(good, product_type if product_type != '(no type)' else '')
        print(f'\n[{t_idx}/{n_types}] type="{product_type}"  ({len(rows)} SKUs)')

        for i in range(0, len(rows), BATCH_SIZE):
            batch = rows[i:i + BATCH_SIZE]
            result = await _enrich_batch(anthropic_client, batch, examples_str, product_type)
            enriched.update(result)
            done_pct = len(enriched) / len(needs) * 100
            elapsed = time.time() - t_start
            print(f'  batch {i // BATCH_SIZE + 1}  enriched so far: {len(enriched)}/{len(needs)} ({done_pct:.0f}%)  {elapsed:.0f}s elapsed', end='\r')

    print(f'\n\nEnrichment complete: {len(enriched)}/{len(needs)} SKUs generated.')

    # Validate — reject results that are shorter than original or suspiciously short
    rejected = []
    for sku, new_st in list(enriched.items()):
        orig_row = df[df['sku'] == sku]
        if orig_row.empty:
            continue
        orig_len = _st_len(orig_row.iloc[0])
        if len(new_st) < max(orig_len, 80):
            rejected.append(sku)
            del enriched[sku]

    if rejected:
        print(f'Rejected {len(rejected)} results (too short or no improvement): {rejected[:5]}...')

    # Update dataframe
    for sku, new_st in enriched.items():
        df.loc[df['sku'] == sku, 'search_text'] = new_st

    enriched_skus = list(enriched.keys())
    _write_xlsx(df, enriched_skus)

    if embed and enriched_skus:
        await _reembed_skus(enriched_skus, df)

    print(f'\nDone. {len(enriched_skus)} SKUs enriched and {"re-embedded" if embed else "written to xlsx (no embed)"}.')


if __name__ == '__main__':
    args = sys.argv[1:]
    do_write = '--write' in args
    do_embed = '--embed' in args
    limit = None
    for a in args:
        if a.startswith('--limit='):
            limit = int(a.split('=', 1)[1])
    asyncio.run(run(write=do_write, embed=do_embed, limit=limit))
